"""
StoreHub Creative Testing Report Template Generator
Creates a Google Sheets-ready .xlsx with all tabs, formulas, and brand formatting.
Data sources: Adveronix (Meta Ads) + Salesforce Connector
Markets: MY + PH | Campaigns: "MY - Creative Testing" / "PH - Creative Testing"
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ── Brand colors ──────────────────────────────────────────────────────────────
ORANGE      = "FF9419"
DARK        = "2F2922"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MID_GRAY    = "D9D9D9"
GREEN_BG    = "E2EFDA"
RED_BG      = "FFCCCC"
YELLOW_BG   = "FFF2CC"
BLUE_BG     = "DDEEFF"

# ── Style helpers ──────────────────────────────────────────────────────────────
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def h1(cell, text=None, bg=DARK, fg=WHITE):
    """Dark header row style."""
    if text is not None:
        cell.value = text
    cell.font = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin()

def h2(cell, text=None, bg=ORANGE, fg=WHITE):
    """Orange sub-header style."""
    if text is not None:
        cell.value = text
    cell.font = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin()

def label(cell, text=None, bold=False, bg=None):
    if text is not None:
        cell.value = text
    cell.font = Font(name="Arial", bold=bold, size=10, color=DARK)
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin()

def val(cell, v=None, fmt=None, bg=None):
    if v is not None:
        cell.value = v
    cell.font = Font(name="Arial", size=10, color=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin()
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    if fmt:
        cell.number_format = fmt

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default blank sheet

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: Helpers
# ══════════════════════════════════════════════════════════════════════════════
hl = wb.create_sheet("Helpers")
hl.sheet_state = "hidden"  # hide from casual viewers

hl["A1"] = "KEY"
hl["B1"] = "VALUE"
hl["C1"] = "NOTES"
for c in ["A1","B1","C1"]: h1(hl[c])

rows = [
    ("THIS_WEEK_START",  '=TODAY()-WEEKDAY(TODAY(),2)+1',         "Monday of current week"),
    ("THIS_WEEK_END",    '=TODAY()-WEEKDAY(TODAY(),2)+7',         "Sunday of current week"),
    ("LAST_WEEK_START",  '=TODAY()-WEEKDAY(TODAY(),2)-6',         "Monday of last week"),
    ("LAST_WEEK_END",    '=TODAY()-WEEKDAY(TODAY(),2)',           "Sunday of last week"),
    ("MY_CAMPAIGN",      "MY - Creative Testing",                 "Meta campaign name filter"),
    ("PH_CAMPAIGN",      "PH - Creative Testing",                 "Meta campaign name filter"),
    ("CT_FILTER",        "*Creative Testing*",                    "Wildcard filter for SUMIFS"),
    ("MY_FILTER",        "MY*",                                   "Wildcard for MY campaigns"),
    ("PH_FILTER",        "PH*",                                   "Wildcard for PH campaigns"),
]
for i, (k, v, n) in enumerate(rows, start=2):
    hl[f"A{i}"] = k;  label(hl[f"A{i}"], bold=True)
    hl[f"B{i}"] = v;  val(hl[f"B{i}"])
    hl[f"C{i}"] = n;  label(hl[f"C{i}"])
    if k in ("THIS_WEEK_START","THIS_WEEK_END","LAST_WEEK_START","LAST_WEEK_END"):
        hl[f"B{i}"].number_format = "DD-MMM-YYYY"

set_col_widths(hl, {"A": 22, "B": 30, "C": 35})

# Named ranges (referenced in formulas as Helpers!B2, etc.)
# Row map: B2=THIS_WEEK_START, B3=THIS_WEEK_END, B4=LAST_WEEK_START, B5=LAST_WEEK_END
# B6=MY_CAMPAIGN, B7=PH_CAMPAIGN, B8=CT_FILTER, B9=MY_FILTER, B10=PH_FILTER

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: RAW - Meta Ads   (populated by Adveronix daily)
# ══════════════════════════════════════════════════════════════════════════════
rm = wb.create_sheet("RAW - Meta Ads")
freeze(rm, "A2")

META_COLS = [
    ("A","Date",              14, "@"),
    ("B","Account Name",      22, "@"),
    ("C","Campaign Name",     30, "@"),
    ("D","Ad Set Name",       25, "@"),
    ("E","Ad Name",           30, "@"),
    ("F","Country",           10, "@"),
    ("G","Spend (MYR)",       14, "#,##0.00"),
    ("H","Impressions",       14, "#,##0"),
    ("I","CPM",               12, "#,##0.00"),
    ("J","Clicks",            10, "#,##0"),
    ("K","CPC",               10, "#,##0.00"),
    ("L","CTR%",              10, "0.00%"),
    ("M","Video Views",       14, "#,##0"),
    ("N","Video View%",       14, "0.00%"),
    ("O","ThruPlays",         12, "#,##0"),
    ("P","Cost per ThruPlay", 18, "#,##0.00"),
]

for col, name, width, fmt in META_COLS:
    h1(rm[f"{col}1"], name)
    rm.column_dimensions[col].width = width

# Instruction row
rm["A2"] = "⬇ Adveronix will populate data below this row. Set filter: Campaign Name contains 'Creative Testing'. Refresh: Daily at 6:00 AM."
rm["A2"].font = Font(name="Arial", italic=True, color="888888", size=9)
rm["A2"].fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
rm.merge_cells("A2:P2")
set_row_height(rm, 1, 35)
set_row_height(rm, 2, 30)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: RAW - Salesforce  (populated by Salesforce Connector daily)
# ══════════════════════════════════════════════════════════════════════════════
rs = wb.create_sheet("RAW - Salesforce")
freeze(rs, "A2")

SF_COLS = [
    ("A","Created Date",    16, "DD-MMM-YYYY"),
    ("B","Lead ID",         20, "@"),
    ("C","Campaign Name",   30, "@"),
    ("D","Ad Set Name",     25, "@"),
    ("E","Ad Name",         30, "@"),
    ("F","Country",         10, "@"),
    ("G","MQL",             10, "@"),
    ("H","SQL",             10, "@"),
    ("I","Won",             10, "@"),
    ("J","Lead Source",     18, "@"),
    ("K","W/C",             14, "DD-MMM-YYYY"),
]

for col, name, width, fmt in SF_COLS:
    h1(rs[f"{col}1"], name)
    rs.column_dimensions[col].width = width

rs["A2"] = "⬇ Salesforce Connector populates below. Required fields: Created Date, Lead ID, Campaign Name, Ad Set Name, Ad Name, Country, MQL (Yes/No), SQL (Yes/No), Won (Yes/No). Filter: Campaign contains 'Creative Testing'. Refresh: Daily."
rs["A2"].font = Font(name="Arial", italic=True, color="888888", size=9)
rs["A2"].fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
rs.merge_cells("A2:K2")
set_row_height(rs, 1, 35)
set_row_height(rs, 2, 30)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4: This Week  (combined MY + PH)
# ══════════════════════════════════════════════════════════════════════════════
tw = wb.create_sheet("This Week")
freeze(tw, "B4")

# Title
tw["A1"] = "STOREHUB — CREATIVE TESTING | THIS WEEK PERFORMANCE"
tw["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
tw["A1"].fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
tw["A1"].alignment = Alignment(horizontal="left", vertical="center")
tw.merge_cells("A1:S1")
set_row_height(tw, 1, 36)

tw["A2"] = '=CONCATENATE("Week: ", TEXT(Helpers!B2,"DD MMM"), " – ", TEXT(Helpers!B3,"DD MMM YYYY"), "   |   Last refresh: ", TEXT(NOW(),"DD-MMM-YYYY HH:MM"))'
tw["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
tw["A2"].alignment = Alignment(horizontal="left", vertical="center")
tw.merge_cells("A2:S2")
set_row_height(tw, 2, 22)

# ── Section A: Summary KPIs ────────────────────────────────────────────────
tw["A3"] = "SUMMARY — MY + PH COMBINED"
tw["A3"].font = Font(name="Arial", bold=True, size=10, color=WHITE)
tw["A3"].fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
tw["A3"].alignment = Alignment(horizontal="left", vertical="center")
tw.merge_cells("A3:S3")
set_row_height(tw, 3, 24)

# KPI header row
kpi_headers = ["METRIC","SPEND","IMPRESSIONS","CPM","CLICKS","CPC","CTR%",
               "VIDEO VIEW%","THRUPLAYS","COST/THRUPLAY","LEADS","MQL","MQL%",
               "SQL","SQL%","WON","CPL","CPMQL","CPSQL","CPWon"]
cols = [get_column_letter(i+1) for i in range(len(kpi_headers))]
for col, hdr in zip(cols, kpi_headers):
    h1(tw[f"{col}4"], hdr)
    tw.column_dimensions[col].width = 13
tw.column_dimensions["A"].width = 22

# Data rows: This Week | Last Week | WoW Change
row_labels = ["This Week", "Last Week", "WoW Change"]
# Shorthand: raw Meta tab = 'RAW - Meta Ads', raw SF tab = 'RAW - Salesforce'
# Column map for RAW - Meta Ads: G=Spend, H=Impressions, I=CPM, J=Clicks, K=CPC, L=CTR%, M=Video Views, N=Video View%, O=ThruPlays, P=Cost per ThruPlay
# Column map for RAW - Salesforce: A=Created Date, G=MQL, H=SQL, I=Won

for r_offset, lbl in enumerate(row_labels):
    row = 5 + r_offset
    set_row_height(tw, row, 22)
    tw[f"A{row}"] = lbl
    label(tw[f"A{row}"], bold=(lbl != "WoW Change"), bg=LIGHT_GRAY if lbl == "WoW Change" else None)

    if lbl == "WoW Change":
        # WoW change row: percentage difference formulas
        tw[f"B{row}"] = '=IFERROR((B5-B6)/ABS(B6),"")'
        tw[f"C{row}"] = '=IFERROR((C5-C6)/ABS(C6),"")'
        tw[f"D{row}"] = '=IFERROR((D5-D6)/ABS(D6),"")'
        tw[f"E{row}"] = '=IFERROR((E5-E6)/ABS(E6),"")'
        tw[f"F{row}"] = '=IFERROR((F5-F6)/ABS(F6),"")'
        tw[f"G{row}"] = '=IFERROR((G5-G6)/ABS(G6),"")'
        tw[f"H{row}"] = '=IFERROR((H5-H6)/ABS(H6),"")'
        tw[f"I{row}"] = '=IFERROR((I5-I6)/ABS(I6),"")'
        tw[f"J{row}"] = '=IFERROR((J5-J6)/ABS(J6),"")'
        tw[f"K{row}"] = '=IFERROR((K5-K6)/ABS(K6),"")'
        tw[f"L{row}"] = '=IFERROR((L5-L6)/ABS(L6),"")'
        tw[f"M{row}"] = '=IFERROR((M5-M6)/ABS(M6),"")'
        tw[f"N{row}"] = '=IFERROR((N5-N6)/ABS(N6),"")'
        tw[f"O{row}"] = '=IFERROR((O5-O6)/ABS(O6),"")'
        tw[f"P{row}"] = '=IFERROR((P5-P6)/ABS(P6),"")'
        tw[f"Q{row}"] = '=IFERROR((Q5-Q6)/ABS(Q6),"")'
        tw[f"R{row}"] = '=IFERROR((R5-R6)/ABS(R6),"")'
        tw[f"S{row}"] = '=IFERROR((S5-S6)/ABS(S6),"")'
        tw[f"T{row}"] = '=IFERROR((T5-T6)/ABS(T6),"")'
        for col_l in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"]:
            tw[f"{col_l}{row}"].number_format = '+0.0%;-0.0%;0%'
            tw[f"{col_l}{row}"].font = Font(name="Arial", size=10)
            tw[f"{col_l}{row}"].alignment = Alignment(horizontal="center", vertical="center")
            tw[f"{col_l}{row}"].border = thin()
            tw[f"{col_l}{row}"].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        continue

    # Date range for this row
    if lbl == "This Week":
        date_start = "Helpers!B2"
        date_end   = "Helpers!B3"
    else:  # Last Week
        date_start = "Helpers!B4"
        date_end   = "Helpers!B5"

    # Spend
    tw[f"B{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!G:G,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"B{row}"].number_format = "#,##0.00"
    # Impressions
    tw[f"C{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!H:H,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"C{row}"].number_format = "#,##0"
    # CPM
    tw[f"D{row}"] = f"=IFERROR(B{row}/C{row}*1000,\"-\")"
    tw[f"D{row}"].number_format = "#,##0.00"
    # Clicks
    tw[f"E{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!J:J,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"E{row}"].number_format = "#,##0"
    # CPC
    tw[f"F{row}"] = f"=IFERROR(B{row}/E{row},\"-\")"
    tw[f"F{row}"].number_format = "#,##0.00"
    # CTR%
    tw[f"G{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!L:L,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end})/COUNTIFS('RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),\"-\")"
    tw[f"G{row}"].number_format = "0.00%"
    # Video View%
    tw[f"H{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!N:N,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end})/COUNTIFS('RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),\"-\")"
    tw[f"H{row}"].number_format = "0.00%"
    # ThruPlays
    tw[f"I{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!O:O,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"I{row}"].number_format = "#,##0"
    # Cost per ThruPlay
    tw[f"J{row}"] = f"=IFERROR(B{row}/I{row},\"-\")"
    tw[f"J{row}"].number_format = "#,##0.00"
    # Leads
    tw[f"K{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8),0)"
    tw[f"K{row}"].number_format = "#,##0"
    # MQL
    tw[f"L{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8,'RAW - Salesforce'!G:G,\"Yes\"),0)"
    tw[f"L{row}"].number_format = "#,##0"
    # MQL%
    tw[f"M{row}"] = f"=IFERROR(L{row}/K{row},\"-\")"
    tw[f"M{row}"].number_format = "0.0%"
    # SQL
    tw[f"N{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8,'RAW - Salesforce'!H:H,\"Yes\"),0)"
    tw[f"N{row}"].number_format = "#,##0"
    # SQL%
    tw[f"O{row}"] = f"=IFERROR(N{row}/L{row},\"-\")"
    tw[f"O{row}"].number_format = "0.0%"
    # Won
    tw[f"P{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8,'RAW - Salesforce'!I:I,\"Yes\"),0)"
    tw[f"P{row}"].number_format = "#,##0"
    # CPL
    tw[f"Q{row}"] = f"=IFERROR(B{row}/K{row},\"-\")"
    tw[f"Q{row}"].number_format = "#,##0.00"
    # CPMQL
    tw[f"R{row}"] = f"=IFERROR(B{row}/L{row},\"-\")"
    tw[f"R{row}"].number_format = "#,##0.00"
    # CPSQL
    tw[f"S{row}"] = f"=IFERROR(B{row}/N{row},\"-\")"
    tw[f"S{row}"].number_format = "#,##0.00"
    # CPWon
    tw[f"T{row}"] = f"=IFERROR(B{row}/P{row},\"-\")"
    tw[f"T{row}"].number_format = "#,##0.00"

    for col_l in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"]:
        cell = tw[f"{col_l}{row}"]
        cell.font = Font(name="Arial", size=10, color=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin()

# ── Section B: Country Breakdown ──────────────────────────────────────────
tw["A9"] = "THIS WEEK — BY COUNTRY"
tw["A9"].font = Font(name="Arial", bold=True, size=10, color=WHITE)
tw["A9"].fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
tw["A9"].alignment = Alignment(horizontal="left", vertical="center")
tw.merge_cells("A9:T9")
set_row_height(tw, 9, 24)

ctry_headers = ["COUNTRY","SPEND","IMPRESSIONS","CPM","CLICKS","CPC","CTR%",
                "VIDEO VIEW%","THRUPLAYS","COST/THRUPLAY","LEADS","MQL","MQL%",
                "SQL","SQL%","WON","CPL","CPMQL","CPSQL","CPWon"]
ctry_cols = [get_column_letter(i+1) for i in range(len(ctry_headers))]
for col, hdr in zip(ctry_cols, ctry_headers):
    h1(tw[f"{col}10"], hdr)

for r_offset, (country, c_filter) in enumerate([("MY", "Helpers!B9"), ("PH", "Helpers!B10")]):
    row = 11 + r_offset
    set_row_height(tw, row, 22)
    tw[f"A{row}"] = country
    label(tw[f"A{row}"], bold=True)

    date_start = "Helpers!B2"
    date_end   = "Helpers!B3"

    tw[f"B{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!G:G,'RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"B{row}"].number_format = "#,##0.00"
    tw[f"C{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!H:H,'RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"C{row}"].number_format = "#,##0"
    tw[f"D{row}"] = f"=IFERROR(B{row}/C{row}*1000,\"-\")"
    tw[f"D{row}"].number_format = "#,##0.00"
    tw[f"E{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!J:J,'RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"E{row}"].number_format = "#,##0"
    tw[f"F{row}"] = f"=IFERROR(B{row}/E{row},\"-\")"
    tw[f"F{row}"].number_format = "#,##0.00"
    tw[f"G{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!L:L,'RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end})/COUNTIFS('RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),\"-\")"
    tw[f"G{row}"].number_format = "0.00%"
    tw[f"H{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!N:N,'RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end})/COUNTIFS('RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),\"-\")"
    tw[f"H{row}"].number_format = "0.00%"
    tw[f"I{row}"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!O:O,'RAW - Meta Ads'!C:C,{c_filter},'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
    tw[f"I{row}"].number_format = "#,##0"
    tw[f"J{row}"] = f"=IFERROR(B{row}/I{row},\"-\")"
    tw[f"J{row}"].number_format = "#,##0.00"
    tw[f"K{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,{c_filter}),0)"
    tw[f"K{row}"].number_format = "#,##0"
    tw[f"L{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,{c_filter},'RAW - Salesforce'!G:G,\"Yes\"),0)"
    tw[f"L{row}"].number_format = "#,##0"
    tw[f"M{row}"] = f"=IFERROR(L{row}/K{row},\"-\")"
    tw[f"M{row}"].number_format = "0.0%"
    tw[f"N{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,{c_filter},'RAW - Salesforce'!H:H,\"Yes\"),0)"
    tw[f"N{row}"].number_format = "#,##0"
    tw[f"O{row}"] = f"=IFERROR(N{row}/L{row},\"-\")"
    tw[f"O{row}"].number_format = "0.0%"
    tw[f"P{row}"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,{c_filter},'RAW - Salesforce'!I:I,\"Yes\"),0)"
    tw[f"P{row}"].number_format = "#,##0"
    tw[f"Q{row}"] = f"=IFERROR(B{row}/K{row},\"-\")"
    tw[f"Q{row}"].number_format = "#,##0.00"
    tw[f"R{row}"] = f"=IFERROR(B{row}/L{row},\"-\")"
    tw[f"R{row}"].number_format = "#,##0.00"
    tw[f"S{row}"] = f"=IFERROR(B{row}/N{row},\"-\")"
    tw[f"S{row}"].number_format = "#,##0.00"
    tw[f"T{row}"] = f"=IFERROR(B{row}/P{row},\"-\")"
    tw[f"T{row}"].number_format = "#,##0.00"

    for col_l in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"]:
        cell = tw[f"{col_l}{row}"]
        cell.font = Font(name="Arial", size=10, color=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5: Last Week (mirrors This Week but uses last week's date range)
# ══════════════════════════════════════════════════════════════════════════════
lw = wb.create_sheet("Last Week")
freeze(lw, "B4")

lw["A1"] = "STOREHUB — CREATIVE TESTING | LAST WEEK PERFORMANCE"
lw["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
lw["A1"].fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
lw["A1"].alignment = Alignment(horizontal="left", vertical="center")
lw.merge_cells("A1:T1")
set_row_height(lw, 1, 36)

lw["A2"] = '=CONCATENATE("Week: ", TEXT(Helpers!B4,"DD MMM"), " – ", TEXT(Helpers!B5,"DD MMM YYYY"), "   |   Last refresh: ", TEXT(NOW(),"DD-MMM-YYYY HH:MM"))'
lw["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
lw["A2"].alignment = Alignment(horizontal="left", vertical="center")
lw.merge_cells("A2:T2")
set_row_height(lw, 2, 22)

lw["A3"] = "SUMMARY — MY + PH COMBINED (LAST WEEK)"
lw["A3"].font = Font(name="Arial", bold=True, size=10, color=WHITE)
lw["A3"].fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
lw["A3"].alignment = Alignment(horizontal="left", vertical="center")
lw.merge_cells("A3:T3")
set_row_height(lw, 3, 24)

for col, hdr in zip(cols, kpi_headers):
    h1(lw[f"{col}4"], hdr)
    lw.column_dimensions[col].width = 13
lw.column_dimensions["A"].width = 22

lw["A5"] = "Last Week"
label(lw["A5"], bold=True)
date_start = "Helpers!B4"
date_end   = "Helpers!B5"
lw[f"B5"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!G:G,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
lw[f"B5"].number_format = "#,##0.00"
lw[f"C5"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!H:H,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
lw[f"C5"].number_format = "#,##0"
lw[f"D5"] = "=IFERROR(B5/C5*1000,\"-\")"
lw[f"D5"].number_format = "#,##0.00"
lw[f"E5"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!J:J,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
lw[f"E5"].number_format = "#,##0"
lw[f"F5"] = "=IFERROR(B5/E5,\"-\")"
lw[f"F5"].number_format = "#,##0.00"
lw[f"G5"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!L:L,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end})/COUNTIFS('RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),\"-\")"
lw[f"G5"].number_format = "0.00%"
lw[f"H5"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!N:N,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end})/COUNTIFS('RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),\"-\")"
lw[f"H5"].number_format = "0.00%"
lw[f"I5"] = f"=IFERROR(SUMIFS('RAW - Meta Ads'!O:O,'RAW - Meta Ads'!C:C,Helpers!B8,'RAW - Meta Ads'!A:A,\">=\"&{date_start},'RAW - Meta Ads'!A:A,\"<=\"&{date_end}),0)"
lw[f"I5"].number_format = "#,##0"
lw[f"J5"] = "=IFERROR(B5/I5,\"-\")"
lw[f"J5"].number_format = "#,##0.00"
lw[f"K5"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8),0)"
lw[f"K5"].number_format = "#,##0"
lw[f"L5"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8,'RAW - Salesforce'!G:G,\"Yes\"),0)"
lw[f"L5"].number_format = "#,##0"
lw[f"M5"] = "=IFERROR(L5/K5,\"-\")"
lw[f"M5"].number_format = "0.0%"
lw[f"N5"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8,'RAW - Salesforce'!H:H,\"Yes\"),0)"
lw[f"N5"].number_format = "#,##0"
lw[f"O5"] = "=IFERROR(N5/L5,\"-\")"
lw[f"O5"].number_format = "0.0%"
lw[f"P5"] = f"=IFERROR(COUNTIFS('RAW - Salesforce'!A:A,\">=\"&{date_start},'RAW - Salesforce'!A:A,\"<=\"&{date_end},'RAW - Salesforce'!C:C,Helpers!B8,'RAW - Salesforce'!I:I,\"Yes\"),0)"
lw[f"P5"].number_format = "#,##0"
lw[f"Q5"] = "=IFERROR(B5/K5,\"-\")"
lw[f"Q5"].number_format = "#,##0.00"
lw[f"R5"] = "=IFERROR(B5/L5,\"-\")"
lw[f"R5"].number_format = "#,##0.00"
lw[f"S5"] = "=IFERROR(B5/N5,\"-\")"
lw[f"S5"].number_format = "#,##0.00"
lw[f"T5"] = "=IFERROR(B5/P5,\"-\")"
lw[f"T5"].number_format = "#,##0.00"
for col_l in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"]:
    cell = lw[f"{col_l}5"]
    cell.font = Font(name="Arial", size=10, color=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 6: Creative Variants
# ══════════════════════════════════════════════════════════════════════════════
cv = wb.create_sheet("Creative Variants")
freeze(cv, "C3")

cv["A1"] = "STOREHUB — CREATIVE TESTING | VARIANT PERFORMANCE TRACKER"
cv["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
cv["A1"].fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
cv["A1"].alignment = Alignment(horizontal="left", vertical="center")
cv.merge_cells("A1:U1")
set_row_height(cv, 1, 36)

cv["A2"] = "Date Range:"
label(cv["A2"], bold=True)
cv["B2"] = '=CONCATENATE(TEXT(Helpers!B2,"DD MMM")," – ",TEXT(TODAY(),"DD MMM YYYY"))'
cv["B2"].font = Font(name="Arial", size=10, color=DARK)
cv.merge_cells("B2:D2")

cv["E2"] = "Market:"
label(cv["E2"], bold=True)
cv["F2"] = "MY + PH"
cv["F2"].font = Font(name="Arial", size=10, color=DARK)

cv["G2"] = "🏆 Winner = lowest CPSQL with spend > 50"
cv["G2"].font = Font(name="Arial", italic=True, size=9, color="888888")
cv.merge_cells("G2:U2")
set_row_height(cv, 2, 22)

# Headers
cv_headers = [
    ("A","AD NAME",           30),
    ("B","THEME",             20),
    ("C","COUNTRY",           10),
    ("D","ITERATION",         12),
    ("E","SPEND",             13),
    ("F","IMPRESSIONS",       14),
    ("G","CPM",               10),
    ("H","CLICKS",            10),
    ("I","CPC",               10),
    ("J","CTR%",              10),
    ("K","VIDEO VIEW%",       13),
    ("L","THRUPLAYS",         12),
    ("M","COST/THRUPLAY",     16),
    ("N","LEADS",             10),
    ("O","MQL",               10),
    ("P","MQL%",              10),
    ("Q","SQL",               10),
    ("R","SQL%",              10),
    ("S","WON",               10),
    ("T","CPSQL",             10),
    ("U","CPWon",             10),
]
for col, hdr, width in cv_headers:
    h1(cv[f"{col}3"], hdr)
    cv.column_dimensions[col].width = width

set_row_height(cv, 3, 30)

# Instructions row
cv["A4"] = "↓ Add ad names from Pencil experiments below. Theme and Country auto-populate if ad name follows convention: [THEME] - [COUNTRY] - [VARIANT]"
cv["A4"].font = Font(name="Arial", italic=True, color="888888", size=9)
cv["A4"].fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
cv.merge_cells("A4:U4")
set_row_height(cv, 4, 28)

# Sample placeholder rows (5 ad variants)
sample_ads = [
    "MY - Creative Testing - Price Anchor v1",
    "MY - Creative Testing - Social Proof v1",
    "MY - Creative Testing - Pain Point v1",
    "PH - Creative Testing - Price Anchor v1",
    "PH - Creative Testing - Social Proof v1",
]
sample_themes = ["Price Anchor", "Social Proof", "Pain Point", "Price Anchor", "Social Proof"]
sample_countries = ["MY", "MY", "MY", "PH", "PH"]
sample_iters = ["Iter 1", "Iter 1", "Iter 1", "Iter 1", "Iter 1"]

for r_offset, (ad, theme, country, iteration) in enumerate(zip(sample_ads, sample_themes, sample_countries, sample_iters)):
    row = 5 + r_offset
    set_row_height(cv, row, 22)

    cv[f"A{row}"] = ad
    label(cv[f"A{row}"])
    cv[f"B{row}"] = theme
    label(cv[f"B{row}"])
    cv[f"C{row}"] = country
    label(cv[f"C{row}"])
    cv[f"D{row}"] = iteration
    label(cv[f"D{row}"])

    # Spend: SUMIFS on RAW - Meta Ads matching exact ad name (column E) and this week
    cv[f"E{row}"] = f'=IFERROR(SUMIFS(\'RAW - Meta Ads\'!G:G,\'RAW - Meta Ads\'!E:E,A{row},\'RAW - Meta Ads\'!A:A,">="&Helpers!B2,\'RAW - Meta Ads\'!A:A,"<="&TODAY()),0)'
    cv[f"E{row}"].number_format = "#,##0.00"
    cv[f"F{row}"] = f'=IFERROR(SUMIFS(\'RAW - Meta Ads\'!H:H,\'RAW - Meta Ads\'!E:E,A{row},\'RAW - Meta Ads\'!A:A,">="&Helpers!B2,\'RAW - Meta Ads\'!A:A,"<="&TODAY()),0)'
    cv[f"F{row}"].number_format = "#,##0"
    cv[f"G{row}"] = f"=IFERROR(E{row}/F{row}*1000,\"-\")"
    cv[f"G{row}"].number_format = "#,##0.00"
    cv[f"H{row}"] = f'=IFERROR(SUMIFS(\'RAW - Meta Ads\'!J:J,\'RAW - Meta Ads\'!E:E,A{row},\'RAW - Meta Ads\'!A:A,">="&Helpers!B2,\'RAW - Meta Ads\'!A:A,"<="&TODAY()),0)'
    cv[f"H{row}"].number_format = "#,##0"
    cv[f"I{row}"] = f"=IFERROR(E{row}/H{row},\"-\")"
    cv[f"I{row}"].number_format = "#,##0.00"
    cv[f"J{row}"] = f'=IFERROR(AVERAGEIFS(\'RAW - Meta Ads\'!L:L,\'RAW - Meta Ads\'!E:E,A{row},\'RAW - Meta Ads\'!A:A,">="&Helpers!B2,\'RAW - Meta Ads\'!A:A,"<="&TODAY()),"-")'
    cv[f"J{row}"].number_format = "0.00%"
    cv[f"K{row}"] = f'=IFERROR(AVERAGEIFS(\'RAW - Meta Ads\'!N:N,\'RAW - Meta Ads\'!E:E,A{row},\'RAW - Meta Ads\'!A:A,">="&Helpers!B2,\'RAW - Meta Ads\'!A:A,"<="&TODAY()),"-")'
    cv[f"K{row}"].number_format = "0.00%"
    cv[f"L{row}"] = f'=IFERROR(SUMIFS(\'RAW - Meta Ads\'!O:O,\'RAW - Meta Ads\'!E:E,A{row},\'RAW - Meta Ads\'!A:A,">="&Helpers!B2,\'RAW - Meta Ads\'!A:A,"<="&TODAY()),0)'
    cv[f"L{row}"].number_format = "#,##0"
    cv[f"M{row}"] = f"=IFERROR(E{row}/L{row},\"-\")"
    cv[f"M{row}"].number_format = "#,##0.00"
    cv[f"N{row}"] = f'=IFERROR(COUNTIFS(\'RAW - Salesforce\'!E:E,A{row},\'RAW - Salesforce\'!A:A,">="&Helpers!B2,\'RAW - Salesforce\'!A:A,"<="&TODAY()),0)'
    cv[f"N{row}"].number_format = "#,##0"
    cv[f"O{row}"] = f'=IFERROR(COUNTIFS(\'RAW - Salesforce\'!E:E,A{row},\'RAW - Salesforce\'!A:A,">="&Helpers!B2,\'RAW - Salesforce\'!A:A,"<="&TODAY(),\'RAW - Salesforce\'!G:G,"Yes"),0)'
    cv[f"O{row}"].number_format = "#,##0"
    cv[f"P{row}"] = f"=IFERROR(O{row}/N{row},\"-\")"
    cv[f"P{row}"].number_format = "0.0%"
    cv[f"Q{row}"] = f'=IFERROR(COUNTIFS(\'RAW - Salesforce\'!E:E,A{row},\'RAW - Salesforce\'!A:A,">="&Helpers!B2,\'RAW - Salesforce\'!A:A,"<="&TODAY(),\'RAW - Salesforce\'!H:H,"Yes"),0)'
    cv[f"Q{row}"].number_format = "#,##0"
    cv[f"R{row}"] = f"=IFERROR(Q{row}/O{row},\"-\")"
    cv[f"R{row}"].number_format = "0.0%"
    cv[f"S{row}"] = f'=IFERROR(COUNTIFS(\'RAW - Salesforce\'!E:E,A{row},\'RAW - Salesforce\'!A:A,">="&Helpers!B2,\'RAW - Salesforce\'!A:A,"<="&TODAY(),\'RAW - Salesforce\'!I:I,"Yes"),0)'
    cv[f"S{row}"].number_format = "#,##0"
    cv[f"T{row}"] = f"=IFERROR(E{row}/Q{row},\"-\")"
    cv[f"T{row}"].number_format = "#,##0.00"
    cv[f"U{row}"] = f"=IFERROR(E{row}/S{row},\"-\")"
    cv[f"U{row}"].number_format = "#,##0.00"

    for col_l in ["E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U"]:
        cell = cv[f"{col_l}{row}"]
        cell.font = Font(name="Arial", size=10, color=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 7: Theme Summary
# ══════════════════════════════════════════════════════════════════════════════
ts = wb.create_sheet("Theme Summary")
freeze(ts, "B3")

ts["A1"] = "STOREHUB — CREATIVE TESTING | THEME PERFORMANCE SUMMARY"
ts["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
ts["A1"].fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
ts["A1"].alignment = Alignment(horizontal="left", vertical="center")
ts.merge_cells("A1:S1")
set_row_height(ts, 1, 36)

ts["A2"] = "Aggregates all variants by theme. Update the theme list in column A as new experiments launch. Formulas aggregate from Creative Variants tab."
ts["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
ts["A2"].alignment = Alignment(horizontal="left", vertical="center")
ts.merge_cells("A2:S2")
set_row_height(ts, 2, 22)

ts_headers = ["THEME","COUNTRY","# VARIANTS","TOTAL SPEND","IMPRESSIONS","CPM","CLICKS","CTR%","THRUPLAYS",
              "LEADS","MQL","MQL%","SQL","SQL%","WON","CPL","CPMQL","CPSQL","CPWon"]
ts_widths   = [20,10,12,14,14,10,10,10,12,10,10,10,10,10,10,10,10,10,10]
ts_cols = [get_column_letter(i+1) for i in range(len(ts_headers))]
for col, hdr, w in zip(ts_cols, ts_headers, ts_widths):
    h1(ts[f"{col}3"], hdr)
    ts.column_dimensions[col].width = w
set_row_height(ts, 3, 30)

sample_themes_all = [
    ("Price Anchor","MY"), ("Social Proof","MY"), ("Pain Point","MY"),
    ("Wildcard / Humour","MY"), ("Price Anchor","PH"), ("Social Proof","PH"),
]
for r_offset, (theme, country) in enumerate(sample_themes_all):
    row = 4 + r_offset
    set_row_height(ts, row, 22)
    ts[f"A{row}"] = theme; label(ts[f"A{row}"], bold=True)
    ts[f"B{row}"] = country; label(ts[f"B{row}"])
    # Count variants: count rows in Creative Variants where col B = theme and col C = country
    ts[f"C{row}"] = f'=COUNTIFS(\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row})'
    ts[f"C{row}"].number_format = "#,##0"
    # Spend
    ts[f"D{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!E:E,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"D{row}"].number_format = "#,##0.00"
    # Impressions
    ts[f"E{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!F:F,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"E{row}"].number_format = "#,##0"
    # CPM
    ts[f"F{row}"] = f"=IFERROR(D{row}/E{row}*1000,\"-\")"
    ts[f"F{row}"].number_format = "#,##0.00"
    # Clicks
    ts[f"G{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!H:H,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"G{row}"].number_format = "#,##0"
    # CTR%
    ts[f"H{row}"] = f"=IFERROR(G{row}/E{row},\"-\")"
    ts[f"H{row}"].number_format = "0.00%"
    # ThruPlays
    ts[f"I{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!L:L,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"I{row}"].number_format = "#,##0"
    # Leads
    ts[f"J{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!N:N,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"J{row}"].number_format = "#,##0"
    # MQL
    ts[f"K{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!O:O,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"K{row}"].number_format = "#,##0"
    # MQL%
    ts[f"L{row}"] = f"=IFERROR(K{row}/J{row},\"-\")"
    ts[f"L{row}"].number_format = "0.0%"
    # SQL
    ts[f"M{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!Q:Q,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"M{row}"].number_format = "#,##0"
    # SQL%
    ts[f"N{row}"] = f"=IFERROR(M{row}/K{row},\"-\")"
    ts[f"N{row}"].number_format = "0.0%"
    # Won
    ts[f"O{row}"] = f'=IFERROR(SUMIFS(\'Creative Variants\'!S:S,\'Creative Variants\'!B:B,A{row},\'Creative Variants\'!C:C,B{row}),0)'
    ts[f"O{row}"].number_format = "#,##0"
    # CPL
    ts[f"P{row}"] = f"=IFERROR(D{row}/J{row},\"-\")"
    ts[f"P{row}"].number_format = "#,##0.00"
    # CPMQL
    ts[f"Q{row}"] = f"=IFERROR(D{row}/K{row},\"-\")"
    ts[f"Q{row}"].number_format = "#,##0.00"
    # CPSQL
    ts[f"R{row}"] = f"=IFERROR(D{row}/M{row},\"-\")"
    ts[f"R{row}"].number_format = "#,##0.00"
    # CPWon
    ts[f"S{row}"] = f"=IFERROR(D{row}/O{row},\"-\")"
    ts[f"S{row}"].number_format = "#,##0.00"

    for col_l in ts_cols[2:]:
        cell = ts[f"{col_l}{row}"]
        cell.font = Font(name="Arial", size=10, color=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin()

# Conditional formatting on CPSQL column (R) — green = low, red = high
ts.conditional_formatting.add(
    f"R4:R{3+len(sample_themes_all)}",
    ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="F8696B"
    )
)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 8: Setup Guide
# ══════════════════════════════════════════════════════════════════════════════
sg = wb.create_sheet("Setup Guide")

sg["A1"] = "SETUP GUIDE — How to Connect Your Data Sources"
sg["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
sg["A1"].fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
sg["A1"].alignment = Alignment(horizontal="left", vertical="center")
sg.merge_cells("A1:C1")
set_row_height(sg, 1, 36)
sg.column_dimensions["A"].width = 5
sg.column_dimensions["B"].width = 35
sg.column_dimensions["C"].width = 65

setup_rows = [
    (None, "━━━ STEP 1: Install Add-ons ━━━", ""),
    ("1", "Install Adveronix",                "Extensions → Add-ons → Get add-ons → search 'Adveronix'. Free tier available."),
    ("2", "Install Salesforce Connector",     "Extensions → Add-ons → Get add-ons → search 'Salesforce Connector for Sheets' by Salesforce."),
    (None, "", ""),
    (None, "━━━ STEP 2: Configure Adveronix (Meta Ads) ━━━", ""),
    ("3", "Open Adveronix sidebar",           "Extensions → Adveronix → Open. Connect your Meta Ads account."),
    ("4", "Set output destination",           "Select sheet: 'RAW - Meta Ads', starting cell A3 (row 1=headers, row 2=instructions)."),
    ("5", "Set account & date range",         "Account: your Meta Ads account. Date Range: Last 30 days (rolling). Refresh: Daily at 6 AM."),
    ("6", "Select fields",                    "Date, Account Name, Campaign Name, Ad Set Name, Ad Name, Country, Amount Spent, Impressions, CPM, Clicks, CPC, CTR (All), Video Plays at 25%, ThruPlay Views, Cost per ThruPlay."),
    ("7", "Set campaign filter",              "Filter: Campaign Name contains 'Creative Testing'  →  this auto-limits to MY and PH creative testing campaigns."),
    (None, "", ""),
    (None, "━━━ STEP 3: Configure Salesforce Connector ━━━", ""),
    ("8", "Open Salesforce Connector",        "Extensions → Salesforce Connector → Open. Connect your Salesforce org."),
    ("9", "Set output destination",           "Select sheet: 'RAW - Salesforce', starting cell A3."),
    ("10","Map fields",                       "Created Date → A, Lead ID → B, Campaign Name (FB Campaign field) → C, Ad Set Name → D, Ad Name → E, Country → F, IsConverted (MQL) → G, IsQualified (SQL) → H, IsWon (Won) → I, LeadSource → J."),
    ("11","Set filter",                       "Filter: Campaign Name contains 'Creative Testing'. Date: Last 30 days rolling."),
    ("12","Set refresh schedule",             "Daily at 7 AM (after Adveronix finishes at 6 AM)."),
    (None, "", ""),
    (None, "━━━ STEP 4: Add Creative Variants ━━━", ""),
    ("13","Go to Creative Variants tab",      "For each Pencil.dev ad variant, add the ad name in column A (must match exactly the Ad Name in Meta Ads)."),
    ("14","Fill in Theme and Country",        "Column B = creative theme (e.g. 'Price Anchor', 'Social Proof'). Column C = MY or PH. Column D = iteration number."),
    ("15","Formulas auto-populate",           "All performance metrics (E onwards) pull automatically from the raw tabs. No manual entry needed after setup."),
    (None, "", ""),
    (None, "━━━ STEP 5: Share with Team ━━━", ""),
    ("16","Share the sheet",                  "File → Share → add team members as Viewers or Commenters. They see live data — no manual refresh needed."),
    ("17","Bookmark This Week tab",           "Set 'This Week' as the default landing tab. Team opens it daily for the morning standup check."),
]

for i, (num, step, detail) in enumerate(setup_rows, start=2):
    set_row_height(sg, i, 28 if detail else 20)
    if num is None and step:
        sg[f"B{i}"] = step
        sg[f"B{i}"].font = Font(name="Arial", bold=True, size=10, color=WHITE)
        sg[f"B{i}"].fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        sg[f"B{i}"].alignment = Alignment(horizontal="left", vertical="center")
        sg.merge_cells(f"B{i}:C{i}")
    elif num:
        sg[f"A{i}"] = num
        sg[f"A{i}"].font = Font(name="Arial", bold=True, size=10, color=WHITE)
        sg[f"A{i}"].fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        sg[f"A{i}"].alignment = Alignment(horizontal="center", vertical="center")
        sg[f"B{i}"] = step
        sg[f"B{i}"].font = Font(name="Arial", bold=True, size=10, color=DARK)
        sg[f"B{i}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sg[f"C{i}"] = detail
        sg[f"C{i}"].font = Font(name="Arial", size=10, color=DARK)
        sg[f"C{i}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
output_path = "/Users/zaidsaad/Downloads/StoreHub_Creative_Testing_Report.xlsx"
wb.save(output_path)
print(f"✅ Report saved to: {output_path}")
print(f"   Tabs: {[s.title for s in wb.worksheets]}")
